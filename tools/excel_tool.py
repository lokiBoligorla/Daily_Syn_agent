import os
import glob
import copy
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def find_excel_file(search_dir: str = ".") -> Optional[str]:
    """
    Dynamically scans the directory (prioritizing 'data/') for a valid .xlsx file.
    Ignores temporary Excel lock files (starting with ~$ ).
    """
    data_dir_pattern = os.path.join(search_dir, "data", "*.xlsx")
    files = glob.glob(data_dir_pattern)
    
    if not files:
        root_dir_pattern = os.path.join(search_dir, "*.xlsx")
        files = glob.glob(root_dir_pattern)
        
    valid_files = [f for f in files if not os.path.basename(f).startswith("~$")]
    
    if valid_files:
        return os.path.abspath(valid_files[0])
    return None

def get_unique_headers(ws) -> List[str]:
    """
    Returns a list of unique headers for every column in the sheet.
    If a header is empty/None, it generates a placeholder like '[Empty Header - Col A]'.
    """
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None and str(val).strip() != "":
            headers.append(str(val).strip())
        else:
            col_letter = get_column_letter(col_idx)
            headers.append(f"[Empty Header - Col {col_letter}]")
    return headers

def get_actual_max_row(ws) -> int:
    """
    Finds the last row containing actual data (non-None, non-whitespace).
    Avoids appending after empty styled rows.
    """
    for r in range(ws.max_row, 0, -1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                return r
    return 1

def get_excel_headers(file_path: str) -> List[str]:
    """
    Loads the excel file and returns the list of unique column headers from the active sheet.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found at: {file_path}")
        
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    
    headers = get_unique_headers(ws)
    
    wb.close()
    return headers

def append_task_rows_to_excel(file_path: str, task_rows: List[Dict[str, Any]]) -> None:
    """
    Appends multiple task rows to the Excel sheet.
    Copies cell formatting (font, border, fill, alignment) from the row above.
    Handles auto-increment of ID/S.No columns if identified.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found at: {file_path}")
        
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 1. Get unique headers and determine actual insertion row
    headers = get_unique_headers(ws)
    actual_max_row = get_actual_max_row(ws)
    
    # We will write after the actual data row
    current_insertion_row = actual_max_row
    
    for row_idx, task_data in enumerate(task_rows):
        current_insertion_row += 1
        
        # We will insert values based on headers
        for col_idx in range(1, len(headers) + 1):
            header_name = headers[col_idx - 1]
            val = task_data.get(header_name, "")
            
            # Check for auto-increment for serial number columns
            header_lower = header_name.lower()
            is_serial = any(kw in header_lower for kw in ["s.no", "sno", "id", "serial", "no.", "seq", "#"])
            
            if is_serial and (val == "INCREMENT" or val == "" or val is None):
                # Try to get previous row's value
                prev_val = ws.cell(row=current_insertion_row - 1, column=col_idx).value
                try:
                    if prev_val is not None:
                        val = int(prev_val) + 1
                    else:
                        val = 1
                except (ValueError, TypeError):
                    val = 1
            
            # Convert date strings to native datetime.date objects for Excel
            if isinstance(val, str) and val.strip() != "":
                from datetime import datetime
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        val = datetime.strptime(val.strip(), fmt).date()
                        break
                    except ValueError:
                        pass
                    
            cell = ws.cell(row=current_insertion_row, column=col_idx, value=val)
            
            # Style preservation: Copy from the previous data row (row 2 or above)
            style_source_row = current_insertion_row - 1 if current_insertion_row > 2 else 2
            if ws.max_row >= style_source_row:
                src_cell = ws.cell(row=style_source_row, column=col_idx)
                
                # Copy font, border, fill, alignment, and number format
                if src_cell.font:
                    cell.font = copy.copy(src_cell.font)
                if src_cell.border:
                    cell.border = copy.copy(src_cell.border)
                if src_cell.fill:
                    cell.fill = copy.copy(src_cell.fill)
                if src_cell.alignment:
                    cell.alignment = copy.copy(src_cell.alignment)
                if src_cell.number_format:
                    cell.number_format = src_cell.number_format
                    
            # Explicit formatting override for date cells to guarantee alignment structure
            if header_name.lower() == "date":
                from openpyxl.styles import Alignment
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "dd/mm/yyyy"
                    
    # Auto-adjust column widths if content is long
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, ws.column_dimensions[col_letter].width or 12)
        
    wb.save(file_path)
    wb.close()
